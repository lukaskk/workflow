from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from .models import  Task
from .forms import  TaskForm
from django.http import FileResponse
import imaplib
import email
import openpyxl
from email import policy
from django.core.files.base import ContentFile
from email.utils import parseaddr
from .forms import UploadFileForm
from .models import Order
from django.core.mail import send_mail
from django.conf import settings

from .models import CustomUser
from .forms import CustomUserCreationForm
from PIL import Image, ImageOps
import os
from django.http import StreamingHttpResponse, HttpResponse
from wsgiref.util import FileWrapper
import openpyxl
from openpyxl.workbook import Workbook
import zipfile
from io import BytesIO
from .models import Order 
import io
from io import BytesIO
import zipfile

import datetime

from  django.shortcuts import render
from .models import Order
from .forms import OrderStatusForm

from django.urls import reverse


from django.shortcuts import render, redirect
from .forms import UpdateOrderForm  # Upewnij się, że masz taki formularz
from .models import Order


from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from .models import Order, OrderPhoto
import zipfile
import openpyxl
from io import BytesIO


from zipfile import ZipFile

from django.http import HttpResponse
from io import BytesIO
import openpyxl
import zipfile
from .models import Order
from django.http import HttpResponse
from io import BytesIO
import zipfile
from app.models import Order  
from django.shortcuts import redirect, get_object_or_404
from .models import Order
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
login_required


def order_photos(request, order_id):
    order = get_object_or_404(Order, pk=order_id)
    return render(request, 'order_photos.html', {'order': order})



def complete_order(request, order_id):
    order = get_object_or_404(Order, pk=order_id)
    order.status = 'Completed'  # Zakładając, że status "Completed" jest prawidłową wartością w modelu
    order.save()
    next_url = request.POST.get('next', '/')
    
    # Przekierowanie do przefiltrowanej tabeli
    return redirect(next_url)




def filter_orders_by_pod(request, inicjaly):
    orders = Order.objects.filter(inicjaly=inicjaly)  # Filtruj zlecenia według nazwy
    total_orders = orders.count()
    completed_orders = orders.filter(status='Completed').count()

    if total_orders > 0:
        progress_percentage = (completed_orders / total_orders) * 100
    else:
        progress_percentage = 0

    context = {
        'orders': orders,
        'filter_name': inicjaly,
        'progress_percentage': progress_percentage,
    }

    return render(request, 'pod.html', context)


 # Upewnij się, że importujesz model Order

def stream_excel_data(orders):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Zrealizowane Zamówienia'
    columns = ["ID Zlecenia", "Nazwa", "Miasto", "Ulica", "Kod Pocztowy", "Data Realizacji", "Status"]
    ws.append(columns)

    for order in orders:
        order_execution_date = order.execution_date.strftime('%Y-%m-%d') if order.execution_date else 'Brak daty'
        ws.append([
            order.order_id, order.name, order.city, order.street,
            order.postal_code, order_execution_date, order.get_status_display()
        ])
    
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file.getvalue()

def stream_zip_data(orders, name):
    in_memory_zip = BytesIO()
    with zipfile.ZipFile(in_memory_zip, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        excel_data = stream_excel_data(orders)
        zip_file.writestr(f"{name}_orders.xlsx", excel_data)

        for order in orders:
            photo_counter = 0
            for photo in order.photos.all():
                photo_counter += 1
                file_path = photo.photo.path
                # Zmiana sposobu tworzenia nazwy folderu i pliku
                folder_name = f"{order.city}_{order.street}".replace(" ", "_")  # Usunięcie spacji
                new_filename = f"{folder_name}/{order.city}_{order.street}_{photo_counter}.jpg".replace(" ", "_")  # Usunięcie spacji
                zip_file.write(file_path, new_filename)
    in_memory_zip.seek(0)
    return in_memory_zip

def download_photos_and_excel(request, name):
    completed_orders = Order.objects.iterator()

    zip_stream = stream_zip_data(completed_orders, name)
    response = StreamingHttpResponse(FileWrapper(zip_stream), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{name}_completed_data.zip"'
    return response




# Upewnij się, że importujesz model OrderPhoto tylko jeśli jest używany gdzie indziej w kodzie

import os
from django.http import HttpResponse
from django.conf import settings
import zipfile
from .models import Order

def download_photos_zip(request, name):
    orders = Order.objects.filter(status='Completed')
    
    photos_per_zip = 100
    current_photo_count = 0
    zip_file_count = 1
    zip_file_paths = []

    # Tworzenie katalogu tymczasowego dla plików ZIP, jeśli nie istnieje
    zip_dir = os.path.join(settings.MEDIA_ROOT, 'temp_zip')
    if not os.path.exists(zip_dir):
        os.makedirs(zip_dir)

    zip_file_path = os.path.join(zip_dir, f"{name}_photos_part{zip_file_count}.zip")
    zip_file = zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED)

    for order in orders:
        for photo in order.photos.all():
            if current_photo_count >= photos_per_zip:
                # Zamykanie bieżącego pliku ZIP
                zip_file.close()
                zip_file_paths.append(zip_file_path)
                
                # Resetowanie licznika i tworzenie nowego pliku ZIP
                zip_file_count += 1
                current_photo_count = 0
                zip_file_path = os.path.join(zip_dir, f"{name}_photos_part{zip_file_count}.zip")
                zip_file = zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED)
            
            city_street_folder = f"{order.city}_{order.street}".replace(" ", "_")
            photo_filename = photo.photo.name.split('/')[-1]
            modified_photo_filename = f"{city_street_folder}_{photo_filename}"
            zip_path = f"{city_street_folder}/{modified_photo_filename}"
            
            photo_file_path = photo.photo.path
            with open(photo_file_path, 'rb') as file_content:
                zip_file.writestr(zip_path, file_content.read())
            
            current_photo_count += 1

    # Dodawanie ostatniego pliku ZIP do listy, jeśli zawiera jakiekolwiek zdjęcia
    if current_photo_count > 0:
        zip_file.close()
        zip_file_paths.append(zip_file_path)

    # Generowanie odpowiedzi z linkami do pobrania plików ZIP
    response_content = "Links to download ZIP files:<br>"
    for path in zip_file_paths:
        file_name = os.path.basename(path)
        download_url = settings.MEDIA_URL + 'temp_zip/' + file_name
        response_content += f'<a href="{download_url}">{file_name}</a><br>'

    return HttpResponse(response_content)





from django.views.generic import ListView, CreateView, UpdateView

from .forms import CustomUserForm
@login_required

def add_user(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('user_list')  # Załóżmy, że user_list to URL do listy użytkowników
    else:
        form = CustomUserCreationForm()
    return render(request, 'add_user.html', {'form': form})
@login_required

def edit_user(request, user_id):
    user = get_object_or_404(CustomUser, pk=user_id)
    if request.method == 'POST':
        form = CustomUserForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            return redirect('user_list')
    else:
        form = CustomUserForm(instance=user)

    return render(request, 'edit_user.html', {'form': form})
@login_required


def delete_user(request, user_id):
    user = get_object_or_404(CustomUser, pk=user_id)
    if request.method == 'POST':
        user.delete()
        return redirect('user_list')

    return render(request, 'delete_user.html', {'user': user})



# Lista klientów
@login_required

# views.py



# views.py
def filter_orders_by_name(request, name):
    orders = Order.objects.filter(name=name)  # Filtruj zlecenia według nazwy
    total_orders = orders.count()
    completed_orders = orders.filter(status='Completed').count()

    if total_orders > 0:
        progress_percentage = (completed_orders / total_orders) * 100
    else:
        progress_percentage = 0

    context = {
        'orders': orders,
        'filter_name': name,
        'progress_percentage': progress_percentage,
    }

    return render(request, 'orders_filtered_list.html', context)



@login_required

def home(request):
    return render(request, 'home.html')

@login_required


def user_list(request):
    users = CustomUser.objects.all()
    return render(request, 'user_list.html', {'users': users})





@login_required  # Upewniamy się, że tylko zalogowani użytkownicy mają dostęp


def task_list(request):
    user = request.user
    tasks = Task.objects.none()  # Domyślnie nie pokazujemy żadnych zadań

    if user.is_authenticated:
        # Sprawdzamy, czy użytkownik jest administratorem lub pracownikiem
        if user.is_administrator() or user.is_employee():
            tasks = Task.objects.all()  # Administrator i pracownik widzą wszystkie zadania
        elif user.is_client():
            # Klient widzi zadania przypisane do jego firmy
            tasks = Task.objects.filter(client_name=user.company)

    return render(request, 'task_list.html', {'tasks': tasks})



@login_required
def create_task(request):
    if request.method == 'POST':
        form = TaskForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('task_list')
    else:
        form = TaskForm()

    return render(request, 'create_task.html', {'form': form})

@login_required
def edit_task(request, task_id):
    task = get_object_or_404(Task, pk=task_id)
    if request.method == 'POST':
        form = TaskForm(request.POST, request.FILES, instance=task)
        if form.is_valid():
            form.save()
            return redirect('task_list')
    else:
        form = TaskForm(instance=task)

    return render(request, 'edit_task.html', {'form': form})

@login_required
def delete_task(request, task_id):
    task = get_object_or_404(Task, pk=task_id)
    if request.method == 'POST':
        task.delete()
        return redirect('task_list')

    return render(request, 'delete_task.html', {'task': task})




import datetime
import random

@login_required


def generate_unique_id():
    # Używamy czasu uniksowego (timestamp) do uzyskania unikalnej wartości
    timestamp = int(datetime.datetime.now().timestamp())
    
    # Generujemy losową liczbę i dodajemy ją do timestamp
    random_number = random.randint(0, 9999)
    unique_id = (timestamp + random_number) % 900000000  # Utrzymanie 6 cyfr

    return str(unique_id).zfill(6)  # Dodajemy zera na początku, jeśli jest krótszy niż 6 cyfr




@login_required

def upload_excel(request):
    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2):
                # Generowanie unikatowego ID
                order_id = row[0].value
                name = row[1].value
                street = row[2].value
                city = row[3].value
                postal_code = row[4].value
                client = row[5].value
                wojewodztwo = row[6].value
                inicjaly = row[7].value
                execution_date = row[8].value
                time_spent = row[9].value

                Order.objects.create(
                    order_id=order_id,
                    name=name,
                    street=street,
                    city=city,
                    postal_code=postal_code,
                    client=client,
                    wojewodztwo=wojewodztwo,
                    inicjaly=inicjaly,
                    execution_date=execution_date,
                    time_spent=time_spent
                )

            return redirect('order_list')

    else:
        form = UploadFileForm()

    return render(request, 'upload_excel.html', {'form': form})

@login_required


def order_list(request):
    orders = Order.objects.all()
    for order in orders:
        for photo in order.photos.all():
            if not photo.photo:
                # Usuwamy rekord zdjęcia, jeśli plik nie istnieje
                photo.delete()
        # Dołącz formularz jako atrybut obiektu zamówienia
        order.form = OrderStatusForm(instance=order)
    return render(request, 'order_list.html', {'orders': orders})


from app.forms import OrderForm # Zakładając, że masz formularz OrderForm
@login_required

def create_order(request):
    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('order_list')  # Przekierowanie do listy zleceń
    else:
        form = OrderForm()
    return render(request, 'create_order.html', {'form': form})
@login_required

def edit_order(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        form = EditOrderForm(request.POST, instance=order)
        photo_form = OrderPhotoForm(request.POST, request.FILES)
        form_valid = form.is_valid()
        photo_form_valid = photo_form.is_valid() and 'photo' in request.FILES  # Sprawdzenie, czy zdjęcie zostało przesłane

        if form_valid and photo_form_valid:
            form.save()
            photo_instance = photo_form.save(commit=False)
            photo_instance.order = order
            photo_instance.save()
            return redirect('filter_orders_by_name', name=order.name)
        elif form_valid and not photo_form_valid:
            form.save()  # Zapisz tylko formularz zamówienia, jeśli formularz zdjęcia jest niepoprawny lub pusty
            return redirect('filter_orders_by_name', name=order.name)

    else:
        form = EditOrderForm(instance=order)
        photo_form = OrderPhotoForm()

    return render(request, 'edit_order.html', {'form': form, 'photo_form': photo_form, 'order': order})
@login_required
def delete_order(request, pk):
    order = get_object_or_404(Order, pk=pk)
    order_name = order.name  # Zapamiętaj nazwę zamówienia
    if request.method == 'POST':
        order.delete()
        return redirect('filter_orders_by_name', name=order_name)  # Przekierowanie do filtrowanej listy zamówień po nazwie
    return render(request, 'delete_order.html', {'order': order})

@login_required
def delete_photo(request, photo_id):
    photo = get_object_or_404(OrderPhoto, id=photo_id)
    order_id = photo.order.id
    order_name = photo.order.name  # Zapamiętaj nazwę zamówienia

    photo.photo.delete(save=False)  # Usuwanie pliku zdjęcia
    photo.delete()  # Usuwanie rekordu zdjęcia

    return redirect('filter_orders_by_name', name=order_name)  # Przekierowanie do filtrowanej listy zamówień po nazwie

from .models import  OrderPhoto
from .forms import EditOrderForm, OrderPhotoForm



from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from app.models import Order
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Rejestruj czcionkę
pdfmetrics.registerFont(TTFont('Lato', 'static/fonts/Afacad-VariableFont_wght.ttf'))
import io
@login_required

def generate_pdf(request, order_id):
    order = Order.objects.get(order_id=order_id)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="order_{order.order_id}.pdf"'

    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    p.setFont("Lato", 16)

    # Wymiary strony A4
    page_width, page_height = A4

    # Zaczynamy od góry strony
    y_position = page_height - 30

    # Dodajemy informacje o zleceniu
    fields = [
        f"Order ID: {order.order_id}",
        f"Name: {order.name}",
        f"Street: {order.street}",
        f"City: {order.city}",
        f"Postal Code: {order.postal_code}",
        f"Client: {order.client}",
        f"Execution Date: {order.execution_date.strftime('%Y-%m-%d') if order.execution_date else 'N/A'}",
        f"Assigned User: {order.assigned_user.username if order.assigned_user else 'N/A'}",
        f"Status: {order.get_status_display()}",
    ]

    for field in fields:
        p.drawString(50, y_position, field)
        y_position -= 20

    # Dodajemy zdjęcia
    photos = order.photos.all()
    for photo in photos:
        if y_position < 150:  # Jeśli jesteśmy blisko końca strony, zaczynamy nową
            p.showPage()
            y_position = page_height - 30

        try:
            img_path = photo.photo.path
            img = ImageReader(img_path)
            p.drawImage(img, 50, y_position - 300, width=400, height=300, preserveAspectRatio=True)
            y_position -= 260  # Zmniejszamy y o wysokość obrazu plus trochę miejsca
        except IOError:
            print("Nie można załadować obrazu: ", img_path)

    p.showPage()
    p.save()

    buffer.seek(0)
    return HttpResponse(buffer.getvalue(), content_type='application/pdf')

@login_required

def download_photo(request, photo_id):
    photo = OrderPhoto.objects.get(id=photo_id)
    file_path = photo.photo.path

    with Image.open(file_path) as img:
        # Określenie maksymalnych wymiarów
        max_size = (2272, 1704)

        # Użycie metody thumbnail do zmiany rozmiaru obrazu
        img.thumbnail(max_size, Image.LANCZOS)

        # Zapisywanie przeskalowanego obrazu do strumienia bajtów w pamięci
        img_byte_arr = io.BytesIO()
        img.save(img_byte_arr, format='JPEG')  # Możesz dostosować format w razie potrzeby
        img_byte_arr.seek(0)

    return FileResponse(img_byte_arr, as_attachment=True, filename=photo.photo.name)



#===================================================


IMAP_SERVER = 'lukaskk.nazwa.pl'
SMTP_SERVER = 'lukaskk.nazwa.pl'
EMAIL_ACCOUNT = 'raport@jklm.eu'
EMAIL_PASSWORD = 'rt4523aaQQ@@!!'
IMAP_PORT = 993







def check_emails():
    mail = imaplib.IMAP4_SSL(IMAP_SERVER)
    mail.login(EMAIL_ACCOUNT, EMAIL_PASSWORD)
    mail.select('inbox')

    result, data = mail.search(None, 'ALL')
    mail_ids = data[0].split()

    for mail_id in mail_ids:
        result, data = mail.fetch(mail_id, '(RFC822)')
        raw_email = data[0][1]
        email_message = email.message_from_bytes(raw_email, policy=policy.default)

        # Analiza tytułu wiadomości e-mail
        email_subject = email_message['subject']
        order_id = email_subject
       
        if order_id:
            try:
                print(f"Order with ID {order_id} dodanet.")
                order = Order.objects.get(order_id=order_id)
                update_order_with_email(order, email_message)
                # Oznaczanie wiadomości jako usuniętej
                mail.store(mail_id, '+FLAGS', '\\Deleted')
            except Order.DoesNotExist:
                print(f"Order with ID {order_id} does not exist.")
        else:
            print("Order ID not found in email subject.")

    # Usunięcie oznaczonych wiadomości
    mail.expunge()
    mail.close()
    mail.logout()
    



def extract_order_id(subject):
    prefix = "Order ID: "
    if subject.startswith(prefix):
        return subject[len(prefix):].strip()
    return None

def update_order_with_email(order, email_message):
    
    order.status = 'Needs Review'
    order.execution_date = datetime.datetime.now()
    order.save()
    sender_email = parseaddr(email_message['From'])[1]  # Pobranie adresu e-mail nadawcy
    photos_saved = []
    for part in email_message.walk():
        if part.get_content_maintype() == 'multipart' or part.get('Content-Disposition') is None:
            continue
        file_data = part.get_payload(decode=True)
        file_name = part.get_filename()

        if file_data and file_name:
            resized_image = resize_image(file_data)
            photo = OrderPhoto(order=order)
            photo.photo.save(file_name, ContentFile(resized_image))
            photos_saved.append(file_name)
           
    #send_feedback_email(order.order_id, f"Zlecenie {order.order_id} zostało zaktualizowane. Miasto: {order.city}, Ulica: {order.street}, Zapisane zdjęcia: {', '.join(photos_saved)}", EMAIL_ACCOUNT, EMAIL_PASSWORD, sender_email)
    return f"Order with ID {order.order_id} has been updated successfully."

def resize_image(file_data, max_width=2016, max_height=3040):
    with Image.open(io.BytesIO(file_data)) as img:
        # Usuwanie kanału alfa dla obrazów RGBA i konwersja do RGB
        if img.mode in ("RGBA", "LA"):
            background = Image.new(img.mode[:-1], img.size, (255, 255, 255))
            background.paste(img, img.split()[-1])  # Usunięcie kanału alfa
            img = background.convert("RGB")
        
        img.thumbnail((max_width, max_height), Image.LANCZOS)
        byte_arr = io.BytesIO()
        img.save(byte_arr, format='JPEG')  # Teraz zapis będzie działał bez błędu
        return byte_arr.getvalue()




def update_database(request):
    check_emails()
    return redirect(reverse('home'))


#===================================================================

from django.shortcuts import redirect, get_object_or_404




from django.shortcuts import get_object_or_404, redirect
from .models import Order

def update_order_status(request, order_id):
    order = get_object_or_404(Order, pk=order_id)
    if request.method == 'POST':
        form = OrderStatusForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            # Przekierowanie z powrotem do filtrowanej listy zleceń, używając nazwy zamówienia
            return redirect('filter_orders_by_name', name=order.name)
    # Tutaj można dodać obsługę błędów formularza lub przekierowanie w przypadku nie-POST



